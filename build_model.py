from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = "LMT_Valuation.xlsx"

BLUE = "0000FF"
BLACK = "000000"
GREEN = "008000"
WHITE = "FFFFFF"
DARK = "1F4E78"
LIGHT_BLUE = "D9EAF7"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"
GRAY = "D9E1F2"

money_fmt = '$#,##0;[Red]($#,##0);-'
per_share_fmt = '$0.00;[Red]($0.00);-'
pct_fmt = '0.0%;[Red](0.0%);-'
multiple_fmt = '0.0x;[Red](0.0x);-'
number_fmt = '#,##0;[Red](#,##0);-'


def style_title(ws, text, last_col):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)


def section(ws, row, text, last_col):
    ws.cell(row, 1, text)
    ws.cell(row, 1).font = Font(bold=True, color=WHITE)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=DARK)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)


def header_row(ws, row, cols):
    fill = PatternFill("solid", fgColor=GRAY)
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def input_cell(cell, fill=False):
    cell.font = Font(color=BLUE)
    if fill:
        cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)


def formula_cell(cell):
    cell.font = Font(color=BLACK)


def link_cell(cell):
    cell.font = Font(color=GREEN)


def add_comment(cell, text):
    from openpyxl.comments import Comment

    cell.comment = Comment(text, "Sid")


def format_range(ws):
    thin = Side(style="thin", color="BFBFBF")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center")
    ws.sheet_view.showGridLines = False


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


wb = Workbook()
wb.remove(wb.active)

summary = wb.create_sheet("Summary")
assump = wb.create_sheet("Assumptions")
dcf = wb.create_sheet("DCF")
comps = wb.create_sheet("Comps")
sens = wb.create_sheet("Sensitivity")

for ws in [summary, assump, dcf, comps, sens]:
    ws.freeze_panes = "B4"

# Assumptions
style_title(assump, "Lockheed Martin Valuation Assumptions", 7)
assump["A2"] = "Currency in $mm except per-share data. Blue font cells are editable hard inputs. Green font cells link across sheets."
assump["A2"].alignment = Alignment(wrap_text=True)
assump.merge_cells("A2:G2")

assump.append(["Group", "Input", "Value", "Units", "Source", "Verify", "Notes"])
header_row(assump, 3, 7)

assumption_rows = [
    ("Historical actuals", "FY2025 revenue", 75048, "$mm", "LMT FY2025 10-K, net sales", "[VERIFY]", "Reported consolidated net sales."),
    ("Historical actuals", "FY2025 operating profit", 7731, "$mm", "LMT FY2025 10-K, operating profit", "[VERIFY]", "Use as EBIT proxy for margin anchor."),
    ("Historical actuals", "FY2025 D&A", 1800, "$mm", "LMT FY2025 10-K, cash flow statement", "[VERIFY]", "Approximate depreciation and amortization."),
    ("Historical actuals", "FY2025 capex", 1850, "$mm", "LMT FY2025 10-K, additions to PP&E", "[VERIFY]", "Approximate capital expenditures."),
    ("Historical actuals", "Cash and equivalents", 3100, "$mm", "LMT FY2025 10-K, balance sheet", "[VERIFY]", "Approximate cash balance."),
    ("Historical actuals", "Total debt", 20000, "$mm", "LMT FY2025 10-K, balance sheet debt", "[VERIFY]", "Approximate short-term plus long-term debt."),
    ("Historical actuals", "Diluted shares", 232.5, "mm shares", "LMT FY2025 10-K, diluted weighted-average shares", "[VERIFY]", "Approximate diluted share count."),
    ("Market", "Current share price", 493.60, "$/share", "MarketWatch close on 2026-06-22", "[VERIFY]", "Recent market price used for upside/downside."),
    ("Market", "Risk-free rate", 0.0449, "%", "10Y U.S. Treasury around 2026-06-24", "[VERIFY]", "Recent 10-year Treasury yield."),
    ("Market", "Equity risk premium", 0.0500, "%", "Analyst assumption", "[VERIFY]", "Midpoint of common U.S. ERP range."),
    ("Market", "Levered beta", 0.63, "x", "Market data provider beta", "[VERIFY]", "Defensive aerospace and defense profile."),
    ("Market", "Pre-tax cost of debt", 0.0480, "%", "Analyst assumption based on credit profile", "[VERIFY]", "Approximate current borrowing cost."),
    ("Capital structure", "Debt weight", 0.22, "%", "Analyst target capital structure", "[VERIFY]", "Target debt as percent of enterprise value."),
    ("Operating drivers", "2026 revenue growth", 0.040, "%", "Analyst forecast based on backlog and FY2026 guide", "[VERIFY]", "Year 1 growth."),
    ("Operating drivers", "2027 revenue growth", 0.035, "%", "Analyst forecast", "[VERIFY]", "Year 2 growth."),
    ("Operating drivers", "2028 revenue growth", 0.030, "%", "Analyst forecast", "[VERIFY]", "Year 3 growth."),
    ("Operating drivers", "2029 revenue growth", 0.027, "%", "Analyst forecast", "[VERIFY]", "Year 4 growth."),
    ("Operating drivers", "2030 revenue growth", 0.025, "%", "Analyst forecast", "[VERIFY]", "Year 5 growth."),
    ("Operating drivers", "2026 EBIT margin", 0.104, "%", "Analyst forecast tied to FY2025 operating margin", "[VERIFY]", "Year 1 margin."),
    ("Operating drivers", "2027 EBIT margin", 0.106, "%", "Analyst forecast", "[VERIFY]", "Year 2 margin."),
    ("Operating drivers", "2028 EBIT margin", 0.107, "%", "Analyst forecast", "[VERIFY]", "Year 3 margin."),
    ("Operating drivers", "2029 EBIT margin", 0.108, "%", "Analyst forecast", "[VERIFY]", "Year 4 margin."),
    ("Operating drivers", "2030 EBIT margin", 0.108, "%", "Analyst forecast", "[VERIFY]", "Year 5 margin."),
    ("Operating drivers", "Cash tax rate", 0.170, "%", "LMT FY2025 10-K effective tax rate and analyst normalization", "[VERIFY]", "Normalized cash tax rate."),
    ("Operating drivers", "D&A % of revenue", 0.024, "%", "LMT FY2025 10-K D&A divided by revenue", "[VERIFY]", "Used for forecast D&A."),
    ("Operating drivers", "Capex % of revenue", 0.025, "%", "LMT FY2025 10-K capex divided by revenue", "[VERIFY]", "Used for forecast capex."),
    ("Operating drivers", "Net working capital % of revenue", 0.050, "%", "Analyst assumption based on balance sheet working capital", "[VERIFY]", "Used to calculate change in NWC."),
    ("Terminal value", "Perpetuity growth", 0.0225, "%", "Analyst assumption", "[VERIFY]", "Long-term defense growth below nominal GDP."),
    ("Terminal value", "Exit EV/EBITDA multiple", 12.0, "x", "Peer trading range", "[VERIFY]", "Cross-check terminal multiple."),
    ("Segment mix", "Aeronautics sales", 30260, "$mm", "LMT FY2025 segment sales", "[VERIFY]", "Segment revenue."),
    ("Segment mix", "Missiles and Fire Control sales", 14450, "$mm", "LMT FY2025 segment sales", "[VERIFY]", "Segment revenue."),
    ("Segment mix", "Rotary and Mission Systems sales", 17310, "$mm", "LMT FY2025 segment sales", "[VERIFY]", "Segment revenue."),
    ("Segment mix", "Space sales", 13030, "$mm", "LMT FY2025 segment sales", "[VERIFY]", "Segment revenue."),
]

for idx, row in enumerate(assumption_rows, start=4):
    assump.append(list(row))
    input_cell(assump.cell(idx, 3), True)
    add_comment(assump.cell(idx, 3), f"{row[4]} {row[5]}. {row[6]}")
    if isinstance(row[2], float) and row[2] < 1:
        assump.cell(idx, 3).number_format = pct_fmt
    elif row[3] == "x":
        assump.cell(idx, 3).number_format = multiple_fmt
    elif row[3] == "$/share":
        assump.cell(idx, 3).number_format = per_share_fmt
    elif row[3] == "$mm":
        assump.cell(idx, 3).number_format = money_fmt
    else:
        assump.cell(idx, 3).number_format = number_fmt

section(assump, 38, "Formula outputs", 7)
formula_rows = [
    ("Formula outputs", "FY2025 EBIT margin", "=C5/C4", "%", "Formula", "", "Operating profit divided by revenue."),
    ("Formula outputs", "Net debt", "=C9-C8", "$mm", "Formula", "", "Total debt less cash."),
    ("Formula outputs", "Equity weight", "=1-C16", "%", "Formula", "", "One less target debt weight."),
    ("Formula outputs", "Cost of equity", "=C12+C14*C13", "%", "Formula", "", "CAPM: risk-free rate plus beta times ERP."),
    ("Formula outputs", "After-tax cost of debt", "=C15*(1-C27)", "%", "Formula", "", "Pre-tax cost of debt after tax shield."),
    ("Formula outputs", "WACC", "=C42*C41+C43*C16", "%", "Formula", "", "Weighted average cost of capital."),
]
for idx, row in enumerate(formula_rows, start=39):
    assump.append(list(row))
    formula_cell(assump.cell(idx, 3))
    if row[3] == "%":
        assump.cell(idx, 3).number_format = pct_fmt
    elif row[3] == "$mm":
        assump.cell(idx, 3).number_format = money_fmt

set_widths(assump, [22, 34, 16, 13, 38, 12, 42])
assump["A2"].font = Font(italic=True)

# DCF
style_title(dcf, "DCF Valuation", 9)
dcf["A2"] = "Five-year unlevered free cash flow forecast with Gordon growth and exit multiple terminal values."
dcf.merge_cells("A2:I2")
dcf.append(["Line item", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E", "Source / logic", "Check"])
header_row(dcf, 3, 9)

dcf_rows = [
    ("Revenue", "='Assumptions'!C4", "=B4*(1+'Assumptions'!C17)", "=C4*(1+'Assumptions'!C18)", "=D4*(1+'Assumptions'!C19)", "=E4*(1+'Assumptions'!C20)", "=F4*(1+'Assumptions'!C21)", "Revenue grows from hard-input drivers", ""),
    ("Revenue growth", "", "=C4/B4-1", "=D4/C4-1", "=E4/D4-1", "=F4/E4-1", "=G4/F4-1", "Forecast assumption", ""),
    ("EBIT margin", "='Assumptions'!C39", "='Assumptions'!C22", "='Assumptions'!C23", "='Assumptions'!C24", "='Assumptions'!C25", "='Assumptions'!C26", "Forecast assumption", ""),
    ("EBIT", "='Assumptions'!C5", "=C4*C6", "=D4*D6", "=E4*E6", "=F4*F6", "=G4*G6", "Revenue times EBIT margin", ""),
    ("Cash taxes", "", "=C7*'Assumptions'!C27", "=D7*'Assumptions'!C27", "=E7*'Assumptions'!C27", "=F7*'Assumptions'!C27", "=G7*'Assumptions'!C27", "EBIT times tax rate", ""),
    ("NOPAT", "", "=C7-C8", "=D7-D8", "=E7-E8", "=F7-F8", "=G7-G8", "EBIT less cash taxes", ""),
    ("D&A", "='Assumptions'!C6", "=C4*'Assumptions'!C28", "=D4*'Assumptions'!C28", "=E4*'Assumptions'!C28", "=F4*'Assumptions'!C28", "=G4*'Assumptions'!C28", "D&A percent of revenue", ""),
    ("Capex", "='Assumptions'!C7", "=C4*'Assumptions'!C29", "=D4*'Assumptions'!C29", "=E4*'Assumptions'!C29", "=F4*'Assumptions'!C29", "=G4*'Assumptions'!C29", "Capex percent of revenue", ""),
    ("Net working capital", "=B4*'Assumptions'!C30", "=C4*'Assumptions'!C30", "=D4*'Assumptions'!C30", "=E4*'Assumptions'!C30", "=F4*'Assumptions'!C30", "=G4*'Assumptions'!C30", "NWC percent of revenue", ""),
    ("Change in NWC", "", "=C12-B12", "=D12-C12", "=E12-D12", "=F12-E12", "=G12-F12", "Period-over-period NWC change", ""),
    ("Unlevered FCF", "", "=C9+C10-C11-C13", "=D9+D10-D11-D13", "=E9+E10-E11-E13", "=F9+F10-F11-F13", "=G9+G10-G11-G13", "NOPAT plus D&A less capex and delta NWC", ""),
    ("WACC", "", "='Assumptions'!C44", "='Assumptions'!C44", "='Assumptions'!C44", "='Assumptions'!C44", "='Assumptions'!C44", "CAPM WACC from Assumptions", ""),
    ("Discount factor", "", "=1/(1+C15)^1", "=1/(1+D15)^2", "=1/(1+E15)^3", "=1/(1+F15)^4", "=1/(1+G15)^5", "Year-end discounting at WACC", ""),
    ("PV of FCF", "", "=C14*C16", "=D14*D16", "=E14*E16", "=F14*F16", "=G14*G16", "FCF times discount factor", ""),
]
for r_idx, row in enumerate(dcf_rows, start=4):
    for c_idx, val in enumerate(row, start=1):
        dcf.cell(r_idx, c_idx, val)
        if isinstance(val, str) and val.startswith("="):
            formula_cell(dcf.cell(r_idx, c_idx))
            if "Assumptions" in val:
                link_cell(dcf.cell(r_idx, c_idx))

section(dcf, 20, "Terminal value and equity bridge", 9)
bridge = [
    ("Terminal FCF", "=G14*(1+'Assumptions'!C31)", "2030E FCF grown at perpetuity rate", "", "", "", "", "", ""),
    ("Terminal value, Gordon growth", "=B21/('Assumptions'!C44-'Assumptions'!C31)", "FCF terminal value", "", "", "", "", "", ""),
    ("PV terminal value, Gordon growth", "=B22*G16", "Discounted at year-5 factor", "", "", "", "", "", ""),
    ("Enterprise value, Gordon growth", "=SUM(C17:G17)+B23", "PV FCF plus PV terminal value", "", "", "", "", "", ""),
    ("Less net debt", "='Assumptions'!C40", "Debt less cash", "", "", "", "", "", ""),
    ("Equity value, Gordon growth", "=B24-B25", "Enterprise value less net debt", "", "", "", "", "", ""),
    ("Diluted shares", "='Assumptions'!C10", "Diluted shares", "", "", "", "", "", ""),
    ("Implied share price, Gordon growth", "=B26/B27", "Equity value divided by shares", "", "", "", "", "", ""),
    ("2030E EBITDA", "=G7+G10", "EBIT plus D&A", "", "", "", "", "", ""),
    ("Terminal value, exit multiple", "=B29*'Assumptions'!C32", "2030E EBITDA times exit multiple", "", "", "", "", "", ""),
    ("PV terminal value, exit multiple", "=B30*G16", "Discounted at year-5 factor", "", "", "", "", "", ""),
    ("Enterprise value, exit multiple", "=SUM(C17:G17)+B31", "PV FCF plus PV terminal value", "", "", "", "", "", ""),
    ("Equity value, exit multiple", "=B32-B25", "Enterprise value less net debt", "", "", "", "", "", ""),
    ("Implied share price, exit multiple", "=B33/B27", "Equity value divided by shares", "", "", "", "", "", ""),
]
for r_idx, row in enumerate(bridge, start=21):
    for c_idx, val in enumerate(row, start=1):
        dcf.cell(r_idx, c_idx, val)
        if isinstance(val, str) and val.startswith("="):
            formula_cell(dcf.cell(r_idx, c_idx))
            if "Assumptions" in val:
                link_cell(dcf.cell(r_idx, c_idx))

section(dcf, 37, "WACC build", 9)
wacc = [
    ("Risk-free rate", "='Assumptions'!C12"),
    ("Levered beta", "='Assumptions'!C14"),
    ("Equity risk premium", "='Assumptions'!C13"),
    ("Cost of equity", "=B38+B39*B40"),
    ("Pre-tax cost of debt", "='Assumptions'!C15"),
    ("Tax rate", "='Assumptions'!C27"),
    ("After-tax cost of debt", "=B42*(1-B43)"),
    ("Debt weight", "='Assumptions'!C16"),
    ("Equity weight", "=1-B45"),
    ("WACC", "=B41*B46+B44*B45"),
]
for idx, row in enumerate(wacc, start=38):
    dcf.cell(idx, 1, row[0])
    dcf.cell(idx, 2, row[1])
    formula_cell(dcf.cell(idx, 2))
    if "Assumptions" in row[1]:
        link_cell(dcf.cell(idx, 2))

for row in [4, 7, 8, 9, 10, 11, 12, 13, 14, 17, 21, 22, 23, 24, 25, 26, 29, 30, 31, 32, 33]:
    for col in range(2, 8):
        dcf.cell(row, col).number_format = money_fmt
for row in [5, 6, 15, 16, 38, 40, 41, 42, 43, 44, 45, 46, 47]:
    for col in range(2, 8):
        dcf.cell(row, col).number_format = pct_fmt
dcf["B39"].number_format = multiple_fmt
dcf["B27"].number_format = number_fmt
dcf["B28"].number_format = per_share_fmt
dcf["B34"].number_format = per_share_fmt
set_widths(dcf, [32, 15, 15, 15, 15, 15, 15, 34, 12])

# Comps
style_title(comps, "Comparable Company Valuation", 11)
comps["A2"] = "Peer inputs are hard-coded market data assumptions flagged for verification. Multiples and implied LMT values are formulas."
comps.merge_cells("A2:K2")
comps.append(["Ticker", "Company", "Share price", "Diluted shares", "Market cap", "Net debt", "Enterprise value", "Revenue", "EBITDA", "Net income", "P/E"])
header_row(comps, 3, 11)
peer_rows = [
    ("RTX", "RTX Corporation", 181.83, 1365.0, "=C4*D4", 32000, "=E4+F4", 84000, 12500, 8000, "=E4/J4"),
    ("NOC", "Northrop Grumman", 507.33, 145.0, "=C5*D5", 13000, "=E5+F5", 41800, 6500, 3800, "=E5/J5"),
    ("GD", "General Dynamics", 295.00, 273.0, "=C6*D6", 11000, "=E6+F6", 48000, 6200, 3900, "=E6/J6"),
    ("LHX", "L3Harris Technologies", 260.00, 188.0, "=C7*D7", 12500, "=E7+F7", 22000, 4200, 2100, "=E7/J7"),
    ("BA", "Boeing", 220.83, 750.0, "=C8*D8", 39000, "=E8+F8", 78000, 4000, -2500, "=IF(J8>0,E8/J8,\"NM\")"),
]
for r_idx, row in enumerate(peer_rows, start=4):
    for c_idx, val in enumerate(row, start=1):
        comps.cell(r_idx, c_idx, val)
        if isinstance(val, str) and val.startswith("="):
            formula_cell(comps.cell(r_idx, c_idx))
        elif c_idx in [3, 4, 6, 8, 9, 10]:
            input_cell(comps.cell(r_idx, c_idx), c_idx in [3, 4, 6, 8, 9, 10])
            add_comment(comps.cell(r_idx, c_idx), "Market or company financial input, [VERIFY] against latest filings or market data.")

comps["L3"] = "EV/Revenue"
comps["M3"] = "EV/EBITDA"
comps["L3"].font = comps["M3"].font = Font(bold=True)
for row in range(4, 9):
    comps.cell(row, 12, f"=G{row}/H{row}")
    comps.cell(row, 13, f"=G{row}/I{row}")
    formula_cell(comps.cell(row, 12))
    formula_cell(comps.cell(row, 13))

section(comps, 11, "LMT implied value from median peer multiples", 13)
implied = [
    ("LMT FY2025 revenue", "='Assumptions'!C4"),
    ("LMT FY2025 EBITDA", "='Assumptions'!C5+'Assumptions'!C6"),
    ("LMT FY2025 NOPAT (net income proxy)", "='Assumptions'!C5*(1-'Assumptions'!C27)"),
    ("Median EV/Revenue", "=MEDIAN(L4:L8)"),
    ("Median EV/EBITDA", "=MEDIAN(M4:M8)"),
    ("Median P/E", "=MEDIAN(K4:K7)"),
    ("Implied EV from revenue", "=B12*B15"),
    ("Implied EV from EBITDA", "=B13*B16"),
    ("Implied equity value from P/E", "=B17*B14"),
    ("Implied price, EV/Revenue", "=(B18-'Assumptions'!C40)/'Assumptions'!C10"),
    ("Implied price, EV/EBITDA", "=(B19-'Assumptions'!C40)/'Assumptions'!C10"),
    ("Implied price, P/E", "=B20/'Assumptions'!C10"),
    ("Comps implied median price", "=MEDIAN(B21:B23)"),
]
for idx, row in enumerate(implied, start=12):
    comps.cell(idx, 1, row[0])
    comps.cell(idx, 2, row[1])
    formula_cell(comps.cell(idx, 2))
    if "Assumptions" in row[1] or "DCF" in row[1]:
        link_cell(comps.cell(idx, 2))

for row in range(4, 9):
    for col in [3]:
        comps.cell(row, col).number_format = per_share_fmt
    for col in [4, 5, 6, 7, 8, 9, 10]:
        comps.cell(row, col).number_format = money_fmt
    comps.cell(row, 11).number_format = multiple_fmt
    comps.cell(row, 12).number_format = multiple_fmt
    comps.cell(row, 13).number_format = multiple_fmt
for row in range(12, 25):
    comps.cell(row, 2).number_format = money_fmt
for row in [15, 16, 17]:
    comps.cell(row, 2).number_format = multiple_fmt
for row in [21, 22, 23, 24]:
    comps.cell(row, 2).number_format = per_share_fmt
set_widths(comps, [12, 28, 14, 16, 16, 14, 18, 14, 14, 14, 12, 14, 14])

# Sensitivity
style_title(sens, "DCF Sensitivity: WACC x Terminal Growth", 8)
sens["A2"] = "Each grid output is a live formula that recalculates implied share price using the row WACC and column terminal growth."
sens.merge_cells("A2:H2")
sens["A4"] = "WACC \\ TGR"
for col, tgr in enumerate([0.015, 0.020, 0.0225, 0.025, 0.030, 0.035], start=2):
    sens.cell(4, col, tgr)
    sens.cell(4, col).number_format = pct_fmt
    input_cell(sens.cell(4, col), True)
for row, wacc_value in enumerate([0.065, 0.070, 0.075, 0.080, 0.085, 0.090], start=5):
    sens.cell(row, 1, wacc_value)
    sens.cell(row, 1).number_format = pct_fmt
    input_cell(sens.cell(row, 1), True)
    for col in range(2, 8):
        col_letter = get_column_letter(col)
        sens.cell(row, col, f"=(('DCF'!C14/(1+$A{row})^1+'DCF'!D14/(1+$A{row})^2+'DCF'!E14/(1+$A{row})^3+'DCF'!F14/(1+$A{row})^4+'DCF'!G14/(1+$A{row})^5)+('DCF'!G14*(1+{col_letter}$4)/($A{row}-{col_letter}$4))/(1+$A{row})^5-'Assumptions'!C40)/'Assumptions'!C10")
        formula_cell(sens.cell(row, col))
        sens.cell(row, col).number_format = per_share_fmt
header_row(sens, 4, 7)
set_widths(sens, [14, 14, 14, 14, 14, 14, 14, 14])

# Summary
style_title(summary, "LMT Valuation Summary", 8)
summary["A2"] = "Investment-banking-style output summary. All valuation outputs are live formulas linked to DCF and Comps."
summary.merge_cells("A2:H2")
summary.append(["Metric", "Low", "Base", "High", "Current", "Upside / Downside", "Method", "Notes"])
header_row(summary, 3, 8)
summary_rows = [
    ("DCF, Gordon growth", "='Sensitivity'!B10", "='DCF'!B28", "='Sensitivity'!G5", "='Assumptions'!C11", "=C4/E4-1", "DCF", "Perpetuity growth method."),
    ("DCF, exit multiple", "=C5*0.90", "='DCF'!B34", "=C5*1.10", "='Assumptions'!C11", "=C5/E5-1", "DCF", "Exit EV/EBITDA method."),
    ("Comparable companies", "=MIN('Comps'!B21:B23)", "='Comps'!B24", "=MAX('Comps'!B21:B23)", "='Assumptions'!C11", "=C6/E6-1", "Comps", "Median peer multiple cross-check."),
    ("Blended target price", "=AVERAGE(B4:B6)", "=AVERAGE(C4:C6)", "=AVERAGE(D4:D6)", "='Assumptions'!C11", "=C7/E7-1", "Blended", "Average of DCF and comps base cases."),
    ("Recommendation", "", '=IF(F7>0.15,"Buy",IF(F7<-0.10,"Sell","Hold"))', "", "", "", "Output", "Formula based on blended upside/downside."),
]
for r_idx, row in enumerate(summary_rows, start=4):
    for c_idx, val in enumerate(row, start=1):
        summary.cell(r_idx, c_idx, val)
        if isinstance(val, str) and val.startswith("="):
            formula_cell(summary.cell(r_idx, c_idx))
            if "Assumptions" in val or "DCF" in val or "Comps" in val or "Sensitivity" in val:
                link_cell(summary.cell(r_idx, c_idx))

section(summary, 11, "Segment mix and risk context", 8)
summary.append(["Segment", "Sales", "Mix", "Notes", "", "", "", ""])
header_row(summary, 12, 4)
segments = [
    ("Aeronautics", "='Assumptions'!C33", "=B13/'Assumptions'!C4", "Largest segment, anchored by F-35."),
    ("Missiles and Fire Control", "='Assumptions'!C34", "=B14/'Assumptions'!C4", "Missile defense and precision fires exposure."),
    ("Rotary and Mission Systems", "='Assumptions'!C35", "=B15/'Assumptions'!C4", "Sikorsky, Aegis, sensors, training."),
    ("Space", "='Assumptions'!C36", "=B16/'Assumptions'!C4", "Satellites, strategic systems, classified programs."),
]
for r_idx, row in enumerate(segments, start=13):
    for c_idx, val in enumerate(row, start=1):
        summary.cell(r_idx, c_idx, val)
        if isinstance(val, str) and val.startswith("="):
            formula_cell(summary.cell(r_idx, c_idx))
            link_cell(summary.cell(r_idx, c_idx))

chart = BarChart()
chart.title = "Implied Share Price Football Field"
chart.y_axis.title = "Price"
chart.x_axis.title = "Method"
data = Reference(summary, min_col=2, max_col=4, min_row=3, max_row=7)
cats = Reference(summary, min_col=1, min_row=4, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 7
chart.width = 16
summary.add_chart(chart, "J3")

for row in range(4, 8):
    for col in range(2, 6):
        summary.cell(row, col).number_format = per_share_fmt
    summary.cell(row, 6).number_format = pct_fmt
summary["C8"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)
summary["C8"].font = Font(bold=True)
for row in range(13, 17):
    summary.cell(row, 2).number_format = money_fmt
    summary.cell(row, 3).number_format = pct_fmt
set_widths(summary, [28, 13, 13, 13, 13, 17, 14, 42, 3, 16, 16, 16])

# Apply global formatting.
for ws in [summary, assump, dcf, comps, sens]:
    format_range(ws)
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.font = Font(color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else BLACK)
            if cell.row == 1:
                cell.border = Border()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

# Reapply important formula/link colors after global border pass.
for ws in [summary, dcf, comps]:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                if any(s in cell.value for s in ["'Assumptions'", "'DCF'", "'Comps'", "'Sensitivity'"]):
                    cell.font = Font(color=GREEN)
                else:
                    cell.font = Font(color=BLACK)

for ws in [assump, comps, sens]:
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.fgColor.rgb in ["00" + LIGHT_YELLOW, LIGHT_YELLOW]:
                cell.font = Font(color=BLUE)

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(OUTPUT)
print(f"Generated {OUTPUT}")
